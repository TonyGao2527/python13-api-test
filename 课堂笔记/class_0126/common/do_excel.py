# @Time     : 2019-03-31 15:28:52
# @Author   : lemon_xiuyu
# @Email    : 5942527@qq.com
# @File     : do_excel.py
# @function : 文件的读写

from openpyxl import load_workbook


# 方法一：类与对象 取值
# 先创建一个类
class Cases:
    """专门存储测试数据"""
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None


# 类里一个参数，读方法里一个参数，分开可同时读取多个表单----芒果老师
class DoExcel:
    def __init__(self, file_name):
        self.file_name = file_name  # 操作的文件
        try:
            self.workbook = load_workbook(self.file_name)  # 实例化一个workbook表单对象
        except FileNotFoundError as e:  # 这里 如文件不存在，应加异常处理--可维护、可扩展，如何做？
            print('断言出错：{}'.format(e))
            print('断言出错：{0} not found,please check file path'.format(file_name))
            raise e

    # 方法一：类与对象 取值
    def read(self, sheet_name):
        """读取数据"""
        self.sheet_name = sheet_name  # 给下面write_back用
        sheet = self.workbook[sheet_name]  # 获取sheet表单
        cases = []  # 列表存所有数据
        for r in range(2, sheet.max_row+1):
            row_data = Cases()
            row_data.case_id = sheet.cell(row=r, column=1).value  # 存储case_id
            row_data.title = sheet.cell(row=r, column=2).value  # 存储title标题
            row_data.url = sheet.cell(row=r, column=3).value  # 存储url地址
            row_data.data = sheet.cell(row=r, column=4).value  # 存储data参数
            row_data.method = sheet.cell(row=r, column=5).value  # 存储method请求
            row_data.expected = sheet.cell(row=r, column=6).value  # 存储expected期望
            if type(row_data.expected) == int:  # 这是干嘛的？？？？？
                row_data.expected = str(row_data.expected)
            cases.append(row_data)  # 将row_data放到cases列表里
        return cases  # for 循环结束返回cases列表

    def write_back(self, row, actual, result):
        """写回数据"""
        sheet = self.workbook[self.sheet_name]  # 获取sheet用read方法里
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.workbook.save(filename=self.file_name)


if __name__ == '__main__':

    # # 类里俩参数，方法里无参数 。这样写死了，只能执行一个sheet表单，要分开
    # # resp = DoExcel('F:\Wenjian\Python_Pycharm\python13-api-test\datas\cases.xlsx', 'login').read()
    # # 直接绝对路径啰嗦 # 类里面两个参数
    # '''
    # from common import contants
    # resp = DoExcel(contants.case_file, 'login').read()  # 引入封装 绝对路径 不啰嗦
    # '''
    # print(resp)

    # 参数分开，类里一个file文件名，方法里一个sheet表名。老师建议
    from common import contants
    do_excel = DoExcel(contants.case_file)
    cases = do_excel.read('login')
    print(cases)  # 读取到login里面所有数据

    from common.request import Request  # 导入Request类，用get、post执行接口
    request = Request()  # 实例化一个对象
    for case in cases:
        print(type(case.data), case.data)  # 打印类型为str
        # 入参处理
        resp = request.request(case.method, case.url, case.data)  # 调用，且拿返回。case.data需转字典，可直接eval()，
        # 但芒果老师建议直接写入request封装类里，避免第一个case转了，第二个case忘记了转。
        print(type(resp.json), resp.json())  # 调用结果为'手机号码不能为空'。因为直接从excel去到的数据是str，调用需要dict
        print(type(resp.text), resp.text)  # str类型，excel里只能写入字符串
        if resp.text == case.expected:
            do_excel.write_back(case.case_id+1, resp.text, 'PASS')
        else:
            do_excel.write_back(case.case_id+1, resp.text, 'FAIL')

        # import json  # 这里的json是库，不是字典格式
        # people = '{"name":"lili","age":18,"married":false,"remark":null}'
        # print(people)  # 字符串json格式
        # o_dict = json.loads(people)  # 通过json库里loads()方法转为字典格式
        # print(type(o_dict), o_dict)
        # # print(eval(people))  # 报错：NameError: name 'false' is not defined
        # # 尝试eval()函数将上面字符串转字典，但不行。
        # # eval()是个表达式，用eval()之前确认里面的数据是python支持的才可以。
        # # 但实战数据难免遇到不支持的，所以要用到json库进行转换。eval()仅限python，json跨语言。so用json
        # # 打印：{"name":"lili","age":18,"married":false,"remark":null}
        # # 打印：<class 'dict'> {'name': 'lili', 'age': 18, 'married': False, 'remark': None} 单双引号，首字母，空
        #
        # # 批量导入数据，100多个字符串入参用json
        # import os  # 需要路径拼接，上面已引入json
        # data_json = os.path.join(contants.data_dir, "data.json")
        # fp = open(data_json)
        # f_dict = json.load(fp=fp)  # 这里时load传文件，不是loads传字符串，传文件流对象。
        # print(f_dict["name"])  # 打印：lili。 说明已转成字典

